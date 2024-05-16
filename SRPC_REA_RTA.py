from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
import streamlit as st
import pandas as pd
import requests
import urllib3
import os
import re
import pdfplumber
import io

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def fetch_pdf_urls(month, year, doc_type):
  base_url = "https://www.srpc.kar.nic.in/website/2023/commercial/"
  pdf_urls = []
  month_locations = {}
  month_location = ["p", "f"]
  for ml in month_location:
    url = f"{base_url}{doc_type.lower()}{month[:3].lower()}{year[-2:]}{ml}.pdf"

    with requests.get(url, verify=False) as response:
      if response.status_code == 200:
        pdf_urls.append(url)
        month_locations[url] = ml
  return pdf_urls, month_locations

def extract_text_from_pdf(pdf_content):
    text = ""
    with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
        for page in pdf.pages:
            text += page.extract_text()
    return text

def find_table(text, search_terms):
      for st in search_terms:
        pattern = re.compile(r'(^.*?\b' + re.escape(st) + r'\b.*?$)', re.IGNORECASE | re.MULTILINE)
        matches = pattern.findall(text)
        for match in matches:
            match_values = match.split()
            if len(match_values) == 2:
                return 2
            elif len(match_values) == 4:
                return 4
            else:
                # Handle other cases as needed
                return None

def extract_data(text, num_columns, search_terms):
  data = []
  date_pattern = r'Actual Meter Reading Available Upto : (\d{4}-\d{2}-\d{2})'
  date_match = re.search(date_pattern, text)
  date_str = date_match.group(1) if date_match else "N/A"

  processed_entities = set()

  for entity in search_terms:
    if entity in processed_entities:
      continue

    if num_columns == 2:
      # entity_pattern = f'(?m)^{entity}\n(.*?)\n'
      entity_pattern = f'(?m)^{entity}\s+(\d+)\s+'
    elif num_columns == 4:
      # entity_pattern = f'(?m)^{entity}\n(.*?)\n(.*?)\n(.*?)\n'
      entity_pattern = f'(?m)^{entity}\s+(.*?)\s+(.*?)\s+(.*?)\s+'
      
    entity_match = re.search(entity_pattern, text, re.MULTILINE)
    if entity_match:
      values = [entity]
      value_lines = entity_match.groups()
      # if num_columns == 2, append empty string for the next two columns
      if num_columns == 2:
        value_lines += ("N/A", ) * 2  # Convert list to tuple here

      values.extend(value_lines[:num_columns -1])  # Extracting data based on num_columns
      # if num_columns == 2 then extend with empty two "N/A" values"
      if num_columns == 2:
        values.extend(["N/A", "N/A"])
      data.append(values)
      processed_entities.add(entity)

  columns = [
      "Entity", "Total Energy Schedule", "Total Actual data",
      "Net Deviation for the purpose of REC"
  ]
  
  # print(f"data = {data}")
  d = pd.DataFrame(data, columns=columns)
  print("d", d)
  return d, date_str

def fetch_data(year, month):
      doc_type = "REA"
      search_terms = ["SPRNG,NPKUNTA", "Fortum Solar,PAVAGADA","SPRNG,PUGULUR", "Sprng Solar India Pvt.Ltd,PAVAGADA"]  # entity names
      solar_entities = {"SPRNG,NPKUNTA", "Fortum Solar,PAVAGADA", "Sprng Solar India Pvt.Ltd,PAVAGADA"}
      non_solar_entities = {"SPRNG,PUGULUR"}
      date_ = ""

      # Fetch PDF URLs
      pdf_urls, month_locations = fetch_pdf_urls(month, year, doc_type)
      solar_data, non_solar_data = [], []

      if pdf_urls:
        st.info("Extracting Data. Please Wait!")
        for pdf_url, month_location in zip(pdf_urls, month_locations.values()):
            with requests.get(pdf_url, verify=False) as response:
                text = extract_text_from_pdf(response.content)
                # print(f"text = {text}")
                num_columns = find_table(text, search_terms)
                # print(f"num_columns = {num_columns}")
                if num_columns:
                    data, date_str = extract_data(text, num_columns, search_terms)
                    date_ = date_str
                    if data is not None:
                        for row in data.values:
                            # print(f"row = {row}")
                            entity = row[0]
                            if entity in solar_entities:
                                solar_data.append((*row, month_location, pdf_url))
                            elif entity in non_solar_entities:
                                non_solar_data.append((*row, month_location, pdf_url))
                            else:
                                print(f"Entity '{entity}' not found in solar_entities or non_solar_entities sets.")
      
        # Define column names
        solar_columns = [
            "Entity", "Total Energy Schedule", "Total Actual data",
            "Net Deviation for the purpose of REC", "Month Location", "PDF URL"
        ]
        non_solar_columns = [
            "Entity", "Total Energy Schedule", "Total Actual data",
            "Net Deviation for the purpose of REC", "Month Location", "PDF URL"
        ]

        # Create DataFrames
        solar_df = pd.DataFrame(solar_data, columns=solar_columns)
        non_solar_df = pd.DataFrame(non_solar_data, columns=non_solar_columns)

        # Save DataFrames as CSV files
        # solar_df.to_csv("solar_data.csv", index=False)
        # non_solar_df.to_csv("non_solar_data.csv", index=False)

        # print("Actual Meter Reading Available Upto :", date_)
        filename = f"Extracted Data_WRPC_SRPC_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
        sheet_name = "SRPC_REA"

        # Check file existence
        if not os.path.exists(filename):
            wb = Workbook()
            wb.save(filename)
        else:
            wb = load_workbook(filename)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        ws2 = wb[sheet_name]

        ws2.cell(row=1, column=1, value="Actual Meter Reading Available Upto :")
        ws2.cell(row=1, column=2, value=date_)

        # Write headers for solar_df
        solar_headers = solar_df.columns
        for col_num, header in enumerate(solar_headers, start=1):
            cell = ws2.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Write solar_df data
        for row_num, (_, row_data) in enumerate(solar_df.iterrows(), start=3):
            for col_num, value in enumerate(row_data, start=1):
                ws2.cell(row=row_num, column=col_num, value=value)

        # Write headers for non_solar_df
        non_solar_headers = non_solar_df.columns
        for col_num, header in enumerate(non_solar_headers, start=1):
            cell = ws2.cell(row=len(solar_df) + 4, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Write non_solar_df data
        for row_num, (_, row_data) in enumerate(non_solar_df.iterrows(), start=len(solar_df) + 5):
            for col_num, value in enumerate(row_data, start=1):
                ws2.cell(row=row_num, column=col_num, value=value)
        # Save the workbook
        wb.save(filename)
        st.success("Data Extracted ✨")
        print("Data Extracted for SRPC_REA ✨")
      else:
        st.error("There's no data available for specified time frame.")
