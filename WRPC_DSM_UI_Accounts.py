from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import BLUE
from datetime import datetime
import pandas as pd
import requests
import os
import streamlit as st
import re
import io  # Import io module
import pdfplumber  # Alternative library for PDF processing
 
selected_pdf = []
search_text = "Arinsun_RUMS"

def create_file(df,fdf, sheet_name):
    filename = f"Extracted Data_WRPC_SRPC_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

    # Check file existence
    if not os.path.exists(filename):
        wb = Workbook()
        wb.save(filename)
    else:
        wb = load_workbook(filename)

    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    ws2 = wb[sheet_name]

    # Write DataFrame column names as headers
    headers = list(df.columns)
    ws2.append(headers)

    # Append data to the worksheet
    for index, row in df.iterrows():  # Iterate over DataFrame rows
        row_list = row.to_list()  # Convert DataFrame row to a list
        ws2.append(row_list)  # Append the row to the worksheet

    ws2.append([])
    ws2.append([])
    text_row = ["(All figures in Rs.)"] + [""] * (len(headers) - 1)
    ws2.append(text_row)

    #FDF
    fheaders = list(fdf.columns)
    ws2.append(fheaders)

    # Append data to the worksheet
    for index, row in fdf.iterrows():  # Iterate over DataFrame rows
        frow_list = row.to_list()  # Convert DataFrame row to a list
        ws2.append(frow_list)  # Append the row to the worksheet

    # Save the workbook
    wb.save(filename)

def create_dataframe(financial_data):
    if not financial_data:
        print("No financial data available.")
        return None

    # Extracted financial data
    entity_name = search_text
    payable = financial_data[0]
    receivable = financial_data[1]
    net_dsm = financial_data[2]
    payable_receivable = financial_data[3]

    # Create DataFrame
    df = pd.DataFrame({
        "Name Of Entity": [entity_name],
        "Payable": [payable],
        "Receivable": [receivable],
        "Net DSM(Rs.)": [net_dsm],
        "Payable/Receivable": [payable_receivable]
    })
    return df

def extract_text_from_pdf(pdf_url):
    try:
        # Download the PDF file
        pdf_data = requests.get(pdf_url).content

        # Load the PDF data
        pdf = pdfplumber.open(io.BytesIO(pdf_data))

        # Extract text from each page
        text = ''
        for page in pdf.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return None

# Function to create clickable links in Excel
def create_hyperlink(url, display_text):
    return f'=HYPERLINK("{url}", "{display_text}")'

def first_row(extracted_text):
    # pattern_combined = re.compile(r'(\d+ \w+_RUMS \d+ \d+ \d+,?\d* \w+)')
    pattern_arinsun = re.compile(r'.*Arinsun_RUMS.*')
    matches = pattern_arinsun.findall(extracted_text)

    if matches:
        return matches[0]
    else:
        print("No match found")

# Function to fetch PDFs
def fetch_pdfs(year, title_filter):
    st.write("Extracting data for WRPC DSM UI Accounts")
    st.warning("Please select files you wants to extract! Then click on  Continue Below")
    wrpc_base_url = "https://www.wrpc.gov.in"
    search_text = "Arinsun_RUMS"
    UI_link = wrpc_base_url + "/assets/data/UI_" + year + '.txt'
    session_state = st.session_state

    if 'checkbox_values' not in session_state:
        session_state.checkbox_values = {}

    response = requests.get(UI_link)

    ui_data = ""
    pdf_links = []
    titles = []
    url_data= {}
    
    if response.status_code == 200:
        ui_data = response.text

    lines = ui_data.split("\n")
    session_state = st.session_state

    for line in lines[1:]:
        parts = line.split(",")
        
        if len(parts) == 5:
            from_date, to_date, link, issue_date, status = parts
            week = re.search(r'week=([\w.]+)', link).group(1)
            yy = re.search(r'yy=(\w+)', link).group(1)
            potential_month = yy.lower()[:3]

            filtered_month = (title_filter.lower().startswith(potential_month) and title_filter) or None

            if filtered_month:
                pdf_link = f"https://www.wrpc.gov.in/htm/{yy}/sum{week}.pdf"
                pdf_links.append(pdf_link)

                status_text = "Revised" if status.strip() == "R" else "Issued"
                output = f"Week{week}: {from_date} to {to_date}\n({status_text} on {issue_date})"
                titles.append(output)

    checkbox_values = []
    url_data = dict(zip(titles, pdf_links))

    checkbox_values = {url: st.checkbox(f"{title}: {url}") for url, title in url_data.items()}

    if st.button("Continue"):
        for url, checked in checkbox_values.items():
            if checked == True:
                selected_pdf.append(url_data[url])

        table_data = []
        first_line_table_data = []

        for pdf_url in selected_pdf:
            print(pdf_url)

            extracted_text = extract_text_from_pdf(pdf_url)
            # Check if extracted text is bytes-like object and decode it to string
            if isinstance(extracted_text, bytes):
                extracted_text = extracted_text.decode("utf-8")


            # break
            pattern_combined = re.compile(r'(\d{2}-\w{3}|Total)\s(Arinsun_RUMS)\s(\d+\.\d+)\s(\d+\.\d+)\s?(\d+\.\d+)?\s?(\d+\.\d+)?\s?(\-?\d+\.\d+)?')
            matches = pattern_combined.findall(extracted_text)
            headers = ['Date', 'Entity', 'Injection', 'Schedule', 'DSM Payable', 'DSM Receivable', 'Net DMC']
            structured_data = []

            for match in matches:
                row_dict = dict(zip(headers, match))
                row_dict['PDF URL'] = create_hyperlink(pdf_url, pdf_url)
                structured_data.append(row_dict)

            table_data.append({})
            table_data.extend(structured_data)

            first_line = first_row(extracted_text)
            # print("First line:", first_line)
            # Split the first_line string into a list of values
            first_line_split = first_line.split()
            f_headers = ['Sr.', 'Name of Entity', 'DSM Charges (Rs.) Payable', 'DSM Charges (Rs.) Receivable', 'Net DSM(Rs.)', 'Net DSM(Rs.) Payable/Receivable' ,]
            first_line_structured_data = []
            frow_dict = dict(zip(f_headers, first_line_split))
            frow_dict['PDF URL'] = create_hyperlink(pdf_url, pdf_url)
            first_line_structured_data.append(frow_dict)
            first_line_table_data.extend(first_line_structured_data)

        df = pd.DataFrame(table_data)
        
        fdf = pd.DataFrame(first_line_table_data)
        # st.write("FDF" ,fdf)

        sheet_name = 'WRPC_DSM'
        create_file(df,fdf, sheet_name)
        st.write("Data extracted for WRPC DSM UI Accounts")
        # st.write(df)
