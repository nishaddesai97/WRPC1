from openpyxl import Workbook, load_workbook
import xml.etree.ElementTree as ET
from openpyxl.styles import Font
from datetime import datetime
import streamlit as st
from io import BytesIO
import pandas as pd
import pdfplumber
import requests
import urllib3
import re
import os

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # Disable SSL certificate verification warning
def get_pdf(url):
    try:
        response = requests.get(url, verify=False)
        response.raise_for_status()  # Raise exception for HTTP errors
        return response.content
    except requests.RequestException as e:
        print(f"Error fetching PDF: {e}")
        return None

def extract_tables_from_pdf_url(pdf_url):
    response = requests.get(pdf_url,  verify=False)
    with pdfplumber.open(BytesIO(response.content)) as pdf:
        all_tables = []
        for page in pdf.pages:
            tables = page.extract_tables()
            all_tables.extend(tables)
    return all_tables

def filter_tables(tables, keywords):
    filtered_tables = []
    for table in tables:
        filtered_table = []
        for row in table:
            for keyword in keywords:
                if keyword in row[0]:
                    filtered_table.append(row)
                    break
        if filtered_table:  # Only include tables with at least one matching row
            filtered_tables.append(filtered_table)
    return filtered_tables

def create_file(DSM_Daywise_St_rows , WS_Seller):
    print("Saving data")
    WS_Seller_cols = ["Entity", "Over Injection Charges (Rs)", "Under Injecton Charges (Rs)",
        "Charges for Drawl without schedule (Rs)", "Final Charges (Rs)", "Payable To Pool/Receviable From Pool" , "PDF Url"]

    DSM_Daywise_St_cols = [ "Station details", "DATE", "Total Schedule (MWHr)",
            "Total Actual (MWHr)", "Total SRAS (MWHr)", "Total Deviation", "Total Overinjection Charges(Rs)", "Total Underinjection Charges (Rs)", "Drawl Charges (Rs)", "PDF Url"]
    
    DSM_Daywise_df = pd.DataFrame(DSM_Daywise_St_rows, columns=DSM_Daywise_St_cols)
    WS_Seller_df = pd.DataFrame(WS_Seller, columns=WS_Seller_cols)

    filename = f"Extracted Data_WRPC_SRPC_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    sheet_name = "SRPC_WA_DSM"

    # Check file existence
    if not os.path.exists(filename):
        wb = Workbook()
        wb.save(filename)
    else:
        wb = load_workbook(filename)
    
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    ws2 = wb[sheet_name]

    ws2.append([])
    # # Write headers 
    ws2.append(DSM_Daywise_St_cols)
    for index, row in DSM_Daywise_df.iterrows():  # Iterate over DataFrame rows
        row_list = row.to_list()  # Convert DataFrame row to a list
        ws2.append(row_list)  # Append the row to the worksheet
    
    ws2.append([])

    # # # Write headers 
    fheaders = list(WS_Seller_cols)
    ws2.append(fheaders)

    # # Write Data
    for index, row in WS_Seller_df.iterrows():  # Iterate over DataFrame rows
        frow_list = row.to_list()  # Convert DataFrame row to a list
        ws2.append(frow_list)
    wb.save(filename)

def create_hyperlink(url, display_text):
    return f'=HYPERLINK("{url}","{display_text}")'

def fetch_data(selected_year, selected_month):
    # Fetching data from the provided URL
    url = "https://www.srpc.kar.nic.in/html/xml-search/data/commercial.xml?cache="
    response = requests.get(url, verify=False)
    if response.status_code == 200:
        # print(f"response.status_code:{response.status_code}")
        xml_data = response.content
        root = ET.fromstring(xml_data)  # Parsing the XML data
        # Store the selected PDF URLs
        selected_urls = []
        st.warning("Please select the week for which you'd like to fetch data, then click 'Continue' below.")

        # Extracting relevant information
        for document in root.findall('document'):
            period = document.find('period').text
            # urls_found = False     # Flag to track if any URLs are found
            if selected_month.lower()[:3] in period.lower() and selected_year in period:
                for i in range(1, 7):  # Assuming URLs are numbered from 1 to 6
                    url_tag = document.find(f'url{i}')
                    if url_tag is not None and url_tag.text.endswith('.pdf') and 'dsm' in url_tag.text.lower():
                        urls_found = True  # Set flag to True if URLs are found
                        pdf_url = url_tag.text
                        title = period
                        # Create a checkbox for each PDF URL title
                        checkbox = st.checkbox(title)
                        # If checkbox is checked, add PDF URL to selected_urls list
                        if checkbox:
                            selected_urls.append(pdf_url)

                # if not urls_found:
                #     st.error("No data were found for the selected period.")

        DSM_Daywise_rows, WS_Seller_rows = [], []
        keywords = ["SPRNG, NPKUNTA", "Sprng, Pugalur", "Fortum Solar", "Sprng Solar India Pvt.Ltd,PAVAGADA"]
        if st.button('Continue'):
            if selected_urls:
                st.info("Extracting data. Please Wait!")
                for url in selected_urls:
                    with requests.get(url, verify=False) as response:
                        tables = extract_tables_from_pdf_url(url)
                        filtered_tables = filter_tables(tables, keywords)
                        for i, table in enumerate(filtered_tables):
                            for row in table:
                                if len(row) < 7:
                                    row.append(create_hyperlink(url,url))
                                    WS_Seller_rows.append(row)
                                else:
                                    row.append(create_hyperlink(url,url))
                                    DSM_Daywise_rows.append(row)

                create_file(DSM_Daywise_rows, WS_Seller_rows)
                st.success("Data extracted ✨")
                print("Extracted SRPC WA DSM✨")
            else:
                st.error("Please select at least one URL before continuing.")
    else:
        st.error("Unable to Fetch the data!")
