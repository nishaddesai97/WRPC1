from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import requests
import pdfplumber  # For PDF processing
import os
import io
import streamlit as st

filename = f"Extracted Data_WRPC_SRPC_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
def create_file(df, sheet_name, pdf_title):
    # Check file existence
    if not os.path.exists(filename):
        wb = Workbook()
    else:
        wb = load_workbook(filename)

    # Remove the default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

    # Check if sheet exists
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    ws2 = wb[sheet_name]

    # Write specific column names as headers
    headers = ["RE Generator", "Schedule (MU)", "Actual (MU)", "Deviation (MU)", "Year", "PDF URL"]
    ws2.append(headers)

    # Append data to the worksheet
    for index, row in df.iterrows():  # Iterate over DataFrame rows
        # Split the 'PDF URL' column into two separate columns for hyperlink function
        row["PDF URL"] = pdf_title
        row_list = row.to_list()  # Convert DataFrame row to a list
        ws2.append(row_list)  # Append the row to the worksheet

    # Save the workbook
    wb.save(filename)

# Function to search for text in PDF and extract the row
def search_text_in_pdf(title, url, search_text):
    # Download the PDF file
    response = requests.get(url)
    pdf_bytes = response.content

    # Open the PDF file
    pdf_document = pdfplumber.open(io.BytesIO(pdf_bytes))

    # Initialize variables to store the extracted row
    found_row = None

    # Loop through each page in the PDF
    for page_num in range(len(pdf_document.pages)):
        page = pdf_document.pages[page_num]
        
        text = page.extract_text()
  
        # Check if the search text exists in the current page
        if search_text in text:
            # Split the text into lines
            lines = text.split("\n")

            # Find the line index containing the search text
            for idx, line in enumerate(lines):
                if search_text in line:
                    # print("search2")
                    # Extract the entire row
                    # found_row = ' '.join(lines[idx:idx]).strip()  # Adjusted to include next 3 lines
                    found_row = lines[idx]
                    print("FOUND ROW ==", found_row)
                    break
            if found_row:
                break  # Exit loop if the first occurrence is found
    pdf_document.close()
    print("found_row", found_row)
    return found_row

# Function to create clickable links in Excel
def create_hyperlink(url, display_text):
    return f'=HYPERLINK("{url}", "{display_text}")'

# Function to convert extracted row to DataFrame
def row_to_dataframe(row, title, url, year):
    if row is None:
        return None

    # Split the row into columns based on whitespace
    columns = row.split()

    # Verify if the number of columns matches the expected number
    if len(columns) % 4 != 0:
        return None

    # Group the columns into chunks of 4 to represent each row of data
    chunked_columns = [columns[i:i+4] for i in range(0, len(columns), 4)]

    # Create DataFrame with specified column names
    df = pd.DataFrame(chunked_columns, columns=["RE Generator", "Schedule (MU)", "Actual (MU)", "Deviation (MU)"])

    # Add year and PDF URL to DataFrame
    df["Year"] = year
    df["PDF URL"] = create_hyperlink(url, title)

    return df

# Function to perform search on multiple PDF URLs and append results into one DataFrame
def search_text_in_multiple_pdfs(pdf_links, search_text, year):
    # print("pdf links", pdf_links)
    # print("search text =", search_text)
    all_rows = []
    for title, url in pdf_links:
        found_row = search_text_in_pdf(title, url, search_text)
        if found_row:
            all_rows.append((found_row, title, url))

    df = pd.concat([row_to_dataframe(row, title, url, year) for row, title, url in all_rows], ignore_index=True)
    return df

# Define a function to extract data based on selected year and title filter
def extract_data(year, title_filter):
    print("Extracting WRPC_Monthly Scheduled Revenue ")
    wrpc_base_url = "https://www.wrpc.gov.in"
    REA_link = f"{wrpc_base_url}/assets/data/REA_{year}.txt"
    response = requests.get(REA_link)
    # print("rea_link", REA_link)
    pdf_title = ""
    pdf_url = ""
    # print(response.status_code)
    if response.status_code == 200:
        rea_data = response.text.split("\n")
        pdf_links = []
        for data_line in rea_data:
            if ".pdf" in data_line:
                data_parts = data_line.split(",")
                if len(data_parts) >= 3:
                    month_year = data_parts[0].strip()
                    title = data_parts[0].strip() + ", " + data_parts[1].strip()
                    pdf_title = title

                    link = wrpc_base_url + "/" + data_parts[2].strip()
                    pdf_url = link
                    if title_filter.lower() in title.lower():
                        pdf_links.append((title, link))
        
        # print("pdf_links>>>>", pdf_links)
        # Search for the text in the multiple PDFs and append the results into one DataFrame
        df = search_text_in_multiple_pdfs(pdf_links, "Arinsun_RUMS", year)  # You can adjust the search text as needed
        
        sheet_name = 'WRPC_Monthly Scheduled Revenue'
        hyper_link = create_hyperlink(pdf_url, pdf_title)
        create_file(df, sheet_name, hyper_link)
        print("Extracted WRPC_Monthly Scheduled Revenue ")
        print(df)
    else:
        print("NO DATA FOUND FOR THE SPECIFIED DATE/ WRPC_Monthly Scheduled Revenue")
        st.warning("There's no data available for WRPC_Monthly Scheduled Revenue in the specified time frame.")
